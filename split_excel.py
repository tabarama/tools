#!/usr/bin/env python3
"""
split_excel_gui.py (оптимизировано + прогресс-бар)

Быстро делит Excel-файл на несколько по значениям выбранного столбца,
сохраняя форматирование/ширины/фильтры/условное форматирование.
Поддержка:
- Имя листа (опционально)
- Номер строки шапки
- Пропуск пустых значений (опция)
- Режим: отдельные файлы или листы в одной книге
- Прогресс-бар и статус
"""
import os
import re
import io
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

class ExcelSplitterApp(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Excel Splitter")
        self.geometry("720x360")
        self.columnconfigure(1, weight=1)

        # Файл
        tk.Label(self, text="Файл Excel:").grid(row=0, column=0, sticky="w", padx=10, pady=5)
        self.file_var = tk.StringVar()
        self.file_entry = tk.Entry(self, textvariable=self.file_var)
        self.file_entry.grid(row=0, column=1, sticky="we", padx=5)
        self.file_entry.bind("<Button-1>", lambda e: self.choose_file())
        tk.Button(self, text="...", width=3, command=self.choose_file).grid(row=0, column=2, padx=(0,10), pady=5)

        # Лист
        tk.Label(self, text="Имя листа (опционально):").grid(row=1, column=0, sticky="w", padx=10, pady=5)
        self.sheet_var = tk.StringVar()
        self.sheet_entry = tk.Entry(self, textvariable=self.sheet_var)
        self.sheet_entry.grid(row=1, column=1, columnspan=2, sticky="we", padx=5)

        # Шапка
        tk.Label(self, text="Номер строки шапки (по умолчанию 1):").grid(row=2, column=0, sticky="w", padx=10, pady=5)
        self.header_var = tk.StringVar(value='1')
        self.header_entry = tk.Entry(self, textvariable=self.header_var)
        self.header_entry.grid(row=2, column=1, columnspan=2, sticky="we", padx=5)

        # Столбец
        tk.Label(self, text="Имя столбца:").grid(row=3, column=0, sticky="w", padx=10, pady=5)
        self.col_var = tk.StringVar()
        self.col_entry = tk.Entry(self, textvariable=self.col_var)
        self.col_entry.grid(row=3, column=1, columnspan=2, sticky="we", padx=5)

        # Опции
        self.skip_empty_var = tk.BooleanVar(value=True)
        tk.Checkbutton(self, text="Пропускать пустые значения (None/пусто)", variable=self.skip_empty_var).grid(row=4, column=1, columnspan=2, sticky="w", padx=5)

        # Режим вывода
        tk.Label(self, text="Режим вывода:").grid(row=5, column=0, sticky="w", padx=10, pady=5)
        self.mode_var = tk.StringVar(value='files')
        mode_frame = tk.Frame(self)
        mode_frame.grid(row=5, column=1, columnspan=2, sticky="w", padx=5)
        tk.Radiobutton(mode_frame, text="Отдельные файлы", variable=self.mode_var, value='files').pack(side='left')
        tk.Radiobutton(mode_frame, text="Листы в одной книге", variable=self.mode_var, value='sheets').pack(side='left', padx=(15, 0))

        # Шорткаты
        for w in (self.file_entry, self.sheet_entry, self.header_entry, self.col_entry):
            self._bind_edit_shortcuts(w)

        # Кнопка Пуск
        self.run_btn = tk.Button(self, text="Запустить", command=self.run_split)
        self.run_btn.grid(row=6, column=1, pady=10)

        # Прогресс-бар и статус
        self.progress = ttk.Progressbar(self, mode='determinate')
        self.progress.grid(row=7, column=0, columnspan=3, sticky="we", padx=10, pady=(5,0))
        self.status_var = tk.StringVar()
        tk.Label(self, textvariable=self.status_var, anchor='w').grid(row=8, column=0, columnspan=3, sticky="we", padx=10, pady=(2,10))

    def _bind_edit_shortcuts(self, widget):
        def handler(event, seq):
            event.widget.event_generate(seq)
            return "break"
        for key, seq in (("<Control-c>", "<<Copy>>"), ("<Control-v>", "<<Paste>>"), ("<Control-x>", "<<Cut>>"), ("<Control-a>", "<<SelectAll>>")):
            widget.bind(key, lambda e, s=seq: handler(e, s))

    def choose_file(self):
        path = filedialog.askopenfilename(
            filetypes=[("Excel", "*.xlsx;*.xlsm")],  # .xls не поддерживается openpyxl
            title="Выберите Excel-файл"
        )
        if path:
            self.file_var.set(path)

    def _delete_rows_and_fix_filter(self, ws, header_row, keep_rows, max_row):
        """Удаляет ненужные строки и обновляет ref AutoFilter."""
        keep_set = set(keep_rows)
        to_delete = [r for r in range(header_row + 1, max_row + 1) if r not in keep_set]
        for a, b in reversed(self._make_ranges(to_delete)):
            ws.delete_rows(a, b - a + 1)
        if ws.auto_filter.ref:
            new_max_row = header_row + len(keep_rows)
            last_col = get_column_letter(ws.max_column)
            ws.auto_filter.ref = f"A{header_row}:{last_col}{new_max_row}"

    @staticmethod
    def _make_ranges(rows):
        if not rows:
            return []
        rows = sorted(rows)
        ranges = []
        start = prev = rows[0]
        for x in rows[1:]:
            if x == prev + 1:
                prev = x
            else:
                ranges.append((start, prev))
                start = prev = x
        ranges.append((start, prev))
        return ranges

    def run_split(self):
        # Блокируем кнопку на время работы
        self.run_btn.config(state='disabled')
        self.status_var.set("")
        self.progress['value'] = 0

        file_path = self.file_var.get().strip()
        sheet = (self.sheet_var.get() or '').strip() or None
        try:
            header_row = int((self.header_var.get() or '1').strip())
            if header_row < 1:
                raise ValueError
        except ValueError:
            messagebox.showerror("Ошибка", "Неверный номер строки шапки.")
            self.run_btn.config(state='normal')
            return
        column = (self.col_var.get() or '').strip()
        if not file_path or not column:
            messagebox.showerror("Ошибка", "Укажите файл, номер строки шапки и имя столбца.")
            self.run_btn.config(state='normal')
            return

        # Читаем файл в память ОДИН раз
        try:
            with open(file_path, 'rb') as f:
                file_bytes = f.read()
        except Exception as e:
            messagebox.showerror("Ошибка чтения", str(e))
            self.run_btn.config(state='normal')
            return

        # База для анализа (из памяти)
        try:
            wb0 = load_workbook(io.BytesIO(file_bytes))
            sheet_name = sheet if sheet else wb0.sheetnames[0]
            ws0 = wb0[sheet_name]
        except Exception as e:
            messagebox.showerror("Ошибка чтения", str(e))
            self.run_btn.config(state='normal')
            return

        # Найдём индекс столбца по шапке
        col_idx = None
        for cell in ws0[header_row]:
            if str(cell.value) == column:
                col_idx = cell.column
                break
        if col_idx is None:
            messagebox.showerror("Ошибка", f"Столбец '{column}' не найден в строке {header_row}.")
            self.run_btn.config(state='normal')
            return

        # Собираем строки по значениям (1 проход по листу)
        from collections import defaultdict
        rows_by_val = defaultdict(list)
        max_row = ws0.max_row
        for r in range(header_row + 1, max_row + 1):
            rows_by_val[ws0.cell(row=r, column=col_idx).value].append(r)

        # Возможно пропустить пустые
        if self.skip_empty_var.get():
            rows_by_val.pop(None, None)
            rows_by_val.pop('', None)

        # Ничего разделять
        if not rows_by_val:
            messagebox.showinfo("Готово", "Нет данных для разделения по заданным условиям.")
            self.run_btn.config(state='normal')
            return

        # Подготовка имён
        directory = os.path.dirname(os.path.abspath(file_path))
        base, ext = os.path.splitext(os.path.basename(file_path))
        ext = ext or '.xlsx'
        def sanitize(name):
            return re.sub(r'[<>:"/\\|?*]', '_', str(name))[:100]
        base_safe = sanitize(base)

        # Прогресс
        keys = sorted(rows_by_val.keys(), key=lambda x: (str(x) if x is not None else ''))
        total = len(keys)
        self.progress['maximum'] = total

        mode = self.mode_var.get()

        if mode == 'sheets':
            self._run_split_sheets(file_bytes, sheet_name, header_row, max_row,
                                   rows_by_val, keys, directory, base_safe, ext, sanitize)
        else:
            self._run_split_files(file_bytes, sheet_name, header_row, max_row,
                                  rows_by_val, keys, directory, base_safe, ext, sanitize)

        self.run_btn.config(state='normal')

    def _run_split_files(self, file_bytes, sheet_name, header_row, max_row,
                         rows_by_val, keys, directory, base_safe, ext, sanitize):
        total = len(keys)
        for i, val in enumerate(keys, 1):
            safe_val = 'None' if val is None else sanitize(val)
            self.status_var.set(f"Обработка: {safe_val} ({i}/{total})")
            self.progress['value'] = i - 1
            self.update_idletasks()
            try:
                wb = load_workbook(io.BytesIO(file_bytes))
                ws = wb[sheet_name]
                keep_rows = rows_by_val[val]
                self._delete_rows_and_fix_filter(ws, header_row, keep_rows, max_row)
                for name in list(wb.sheetnames):
                    if name != sheet_name:
                        wb.remove(wb[name])
                out_path = os.path.join(directory, f"{base_safe}_{safe_val}{ext}")
                wb.save(out_path)
            except Exception as e:
                messagebox.showerror("Ошибка записи", f"{val}: {e}")
            finally:
                self.progress['value'] = i
                self.update_idletasks()

        self.status_var.set("Готово")
        messagebox.showinfo("Готово", "Файлы успешно созданы!")

    def _run_split_sheets(self, file_bytes, sheet_name, header_row, max_row,
                          rows_by_val, keys, directory, base_safe, ext, sanitize):
        total = len(keys)
        try:
            wb = load_workbook(io.BytesIO(file_bytes))
            ws_orig = wb[sheet_name]

            for i, val in enumerate(keys, 1):
                safe_val = 'None' if val is None else sanitize(val)
                self.status_var.set(f"Обработка: {safe_val} ({i}/{total})")
                self.progress['value'] = i - 1
                self.update_idletasks()

                ws_copy = wb.copy_worksheet(ws_orig)
                # Имя листа в Excel ограничено 31 символом
                ws_copy.title = re.sub(r'[<>:"/\\|?*\[\]]', '_', str(safe_val))[:31]
                keep_rows = rows_by_val[val]
                self._delete_rows_and_fix_filter(ws_copy, header_row, keep_rows, max_row)

                self.progress['value'] = i
                self.update_idletasks()

            wb.remove(ws_orig)
            out_path = os.path.join(directory, f"{base_safe}_split{ext}")
            wb.save(out_path)

        except Exception as e:
            messagebox.showerror("Ошибка", str(e))
            return

        self.status_var.set("Готово")
        messagebox.showinfo("Готово", f"Файл сохранён:\n{out_path}")

if __name__ == '__main__':
    app = ExcelSplitterApp()
    app.mainloop()
